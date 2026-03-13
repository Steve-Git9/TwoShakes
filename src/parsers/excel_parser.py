"""Excel parser — handles merged cells, multi-row headers, multiple sheets."""
import os, logging
from typing import Optional
import pandas as pd
import openpyxl
from src.models.schemas import FileMetadata

logger = logging.getLogger(__name__)

def parse_excel(file_path: str, sheet_name: Optional[str] = None) -> tuple[pd.DataFrame, FileMetadata]:
    """Parse an Excel file. sheet_name=None uses first sheet."""
    warnings: list[str] = []
    size_bytes = os.path.getsize(file_path)
    original_filename = os.path.basename(file_path)
    sheet_names: list[str] = []
    try:
        xl = pd.ExcelFile(file_path, engine="openpyxl")
        sheet_names = xl.sheet_names
        target_sheet = sheet_name if sheet_name is not None else sheet_names[0]

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[target_sheet]

        # Unmerge and forward-fill merged cells
        for mr in list(ws.merged_cells.ranges):
            val = ws.cell(mr.min_row, mr.min_col).value
            ws.unmerge_cells(str(mr))
            for row in ws.iter_rows(min_row=mr.min_row, max_row=mr.max_row,
                                    min_col=mr.min_col, max_col=mr.max_col):
                for cell in row:
                    cell.value = val

        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()

        if not data:
            warnings.append("Sheet is empty")
            df = pd.DataFrame()
        else:
            row0 = data[0]
            none_ratio = sum(1 for v in row0 if v is None) / max(len(row0), 1)
            if len(data) > 1 and none_ratio > 0.3:
                row1 = data[1]
                cols = []
                for a, b in zip(row0, row1):
                    parts = [str(v).strip() for v in (a, b) if v is not None and str(v).strip()]
                    cols.append("_".join(parts) if parts else "")
                df = pd.DataFrame(data[2:], columns=cols)
                warnings.append("Multi-row header detected; merged first two rows as column names")
            else:
                df = pd.DataFrame(data[1:], columns=row0)

            df.columns = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(df.columns)]
            df = df.astype(object).where(df.notna(), other=None).astype(str).replace("None", "")

        logger.info(f"Parsed Excel {target_sheet!r}: {len(df)} rows x {len(df.columns)} cols")
    except Exception as e:
        warnings.append(f"Excel parse error: {e}")
        logger.error(f"Failed to parse {file_path}: {e}")
        df = pd.DataFrame()

    return df, FileMetadata(
        original_filename=original_filename, file_format="xlsx",
        size_bytes=size_bytes, sheet_names=sheet_names,
        row_count=len(df), col_count=len(df.columns), parse_warnings=warnings,
    )
